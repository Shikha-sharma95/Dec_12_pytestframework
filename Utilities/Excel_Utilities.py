import openpyxl


def read_data_from_excel(file_path, sheet_name, row_no, col_no):
    excel_file = openpyxl.load_workbook(file_path)  # given file to program
    sheet = excel_file[sheet_name]  # given file's  sheet to program
    data = sheet.cell(row=row_no, column=col_no).value
    return data

def write_data_to_excel(file_path, sheet_name, row_no, col_no, data):
    excel_file = openpyxl.load_workbook(file_path)  # given file to program
    sheet = excel_file[sheet_name]  # given file's  sheet to program
    sheet.cell(row=row_no, column=col_no).value = data #
    excel_file.save(file_path)  # save the changes
    excel_file.close() # close the file

def get_row_count(file_path, sheet_name):
    excel_file = openpyxl.load_workbook(file_path)  # given file to program
    sheet = excel_file[sheet_name]  # given file's  sheet to program
    return sheet.max_row
